import pandas as pd
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Buka dialog pemilihan file
Tk().withdraw()
print("🔍 Silakan pilih file CSV input...")
input_file = askopenfilename(filetypes=[("CSV Files", "*.csv")])

if not input_file:
    print("❌ Tidak ada file yang dipilih.")
    exit()

print(f"📂 File dipilih: {input_file}")

try:
    # Baca file CSV
    df = pd.read_csv(input_file)
except Exception as e:
    print(f"❌ Gagal membaca file: {e}")
    exit()

# Bersihkan nama kolom
df.columns = df.columns.str.strip().str.replace('\u200b', '')

# Validasi kolom yang dibutuhkan
# Tambahkan kolom 'Sign-in error code' jika ada di input
columns_needed = ["Date (UTC)", "User", "User type", "IP address", "Location", "Status", "Application"]
if "Sign-in error code" in df.columns:
    columns_needed.append("Sign-in error code")
missing = [col for col in columns_needed if col not in df.columns]

if missing:
    print(f"⚠️ Kolom berikut tidak ditemukan dalam file: {missing}")
    exit()

# Filter dan proses data
df_filtered = df[columns_needed].copy()
# Filter lokasi yang mengandung 'ID' dan yang nilainya ', ,'
df_filtered = df_filtered[~df_filtered["Location"].astype(str).str.contains("ID", case=False, na=False)]
df_filtered = df_filtered[df_filtered["Location"].astype(str).str.strip() != ", ,"]

# Konversi waktu ke WIB (UTC+7)
df_filtered["Date (UTC)"] = pd.to_datetime(df_filtered["Date (UTC)"], errors='coerce') + pd.Timedelta(hours=7)
df_filtered = df_filtered.dropna(subset=["Date (UTC)"])
df_filtered = df_filtered.sort_values(by="Date (UTC)")
df_filtered["Jam"] = df_filtered["Date (UTC)"].dt.strftime("%H.%M.%S")
df_filtered["Date (UTC)"] = df_filtered["Date (UTC)"].dt.tz_localize(None)

# Siapkan nama file output
now_wib = datetime.utcnow() + timedelta(hours=7)
now_str = now_wib.strftime("%Y-%m-%d_%H-%M-%S")
output_filename = f"entra.PT.smi_{now_str}.xlsx"
output_path = os.path.join(os.path.dirname(input_file), output_filename)

# Simpan awal ke Excel
df_filtered.to_excel(output_path, index=False, sheet_name="Data")

# Buka kembali file untuk formatting
wb = load_workbook(output_path)
ws = wb.active

# Buat tabel dengan gaya
max_row = ws.max_row
max_col = ws.max_column
last_col_letter = get_column_letter(max_col)
table_ref = f"A1:{last_col_letter}{max_row}"

table = Table(displayName="TabelData", ref=table_ref)
style = TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)
table.tableStyleInfo = style
ws.add_table(table)

# Auto-width kolom
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells if cell.value)
    col_letter = column_cells[0].column_letter
    ws.column_dimensions[col_letter].width = length + 2


# Tambahkan ringkasan status login
status_counts = df_filtered['Status'].value_counts()

ws.append([])
ws.append(["Summary Status"])
ws.append(["Success", status_counts.get("Success", 0)])
ws.append(["Failure", status_counts.get("Failure", 0)])
ws.append(["Interrupted", status_counts.get("Interrupted", 0)])
    
# Buat sheet baru untuk template summary
summary_text = f"""Dear Tim IT Security,
Berikut adalah nama user yang terdeteksi login dari luar negeri, {now_wib.strftime('%d %B %Y %H:%M')} WIB, dengan detail berikut:
* Success : {status_counts.get('Success', 0)}
* Failure: {status_counts.get('Failure', 0)}
* Interrupted: {status_counts.get('Interrupted', 0)}
Terima Kasih
"""
ws_summary = wb.create_sheet(title="Summary")
ws_summary["A1"] = summary_text

# Simpan file akhir
wb.save(output_path)



# Tabel Sign-in error code
error_codes = [
    ["Error Code", "Description"],
    ["0", "Success"],
    ["50053", "Account locked"],
    ["50126", "Invalid credentials"],
    ["50125", "Account disabled"],
    ["50076", "MFA required"],
    ["50074", "MFA registration required"],
    ["50055", "Password expired"],
    ["50072", "User password reset required"],
    ["50079", "Session MFA required"],
    ["50144", "Guest user not allowed"],
    ["53003", "Conditional Access policy"],
    ["700016", "Application not found"],
    ["7000215", "Invalid client secret"],
    ["7000225", "Invalid client certificate"],
    ["700082", "SAML assertion error"],
]
# Tambahkan sheet baru ke file Excel untuk error code
ws_error = wb.create_sheet(title="Sign-in Error Code")
for row in error_codes:
    ws_error.append(row)

print(f"✅ File berhasil disimpan dengan nama: {output_path}")
